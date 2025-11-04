"""Excel file parser."""

from pathlib import Path
from .base import BaseParser, ParseResult


class XLSXParser(BaseParser):
    """Excel 파일 파서 (openpyxl 사용)"""

    def parse(self, file_path: Path) -> ParseResult:
        """
        XLSX 파일에서 텍스트를 추출합니다.

        Args:
            file_path: 파싱할 XLSX 파일 경로

        Returns:
            ParseResult: 파싱된 텍스트와 메타데이터
        """
        if not file_path.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

        try:
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise ImportError(
                    "openpyxl이 설치되지 않았습니다. 'pip install openpyxl'로 설치하세요."
                )

            # data_only=True로 수식 대신 값을 읽음
            workbook = load_workbook(file_path, data_only=True)

            # 시트별로 데이터 추출
            sheets_text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_content = [f"=== Sheet: {sheet_name} ==="]

                # 모든 행 읽기
                rows_data = []
                for row in sheet.iter_rows(values_only=True):
                    # 빈 행 건너뛰기
                    if all(cell is None or str(cell).strip() == '' for cell in row):
                        continue

                    # 셀 값을 문자열로 변환
                    row_text = ' | '.join(
                        str(cell) if cell is not None else '' for cell in row
                    )
                    rows_data.append(row_text)

                if rows_data:
                    sheet_content.extend(rows_data)
                else:
                    sheet_content.append("(empty sheet)")

                sheets_text.append('\n'.join(sheet_content))

            content = '\n\n'.join(sheets_text)

            metadata = {
                'file_type': 'xlsx',
                'num_sheets': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'file_size': file_path.stat().st_size,
            }

            # 활성 시트 정보
            if workbook.active:
                metadata['active_sheet'] = workbook.active.title

            return ParseResult(content=content, metadata=metadata)

        except Exception as e:
            raise ValueError(f"XLSX 파싱 실패: {file_path}") from e

    def supports(self, file_path: Path) -> bool:
        """XLSX 파일 확장자 지원 확인"""
        return file_path.suffix.lower() in self.supported_extensions

    @property
    def supported_extensions(self) -> list[str]:
        """지원하는 Excel 파일 확장자"""
        return ['.xlsx', '.xlsm']
